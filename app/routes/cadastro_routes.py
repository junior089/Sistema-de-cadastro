from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file, make_response
from flask_login import login_required, current_user
from app import db
from app.models.cadastro import Cadastro
from app.models.user import User
from app.models.municipio import Municipio
from app.models.estado import Estado
from app.models.descricao import Descricao
from app.models.instituicao import Instituicao
from app.models.atendente import Atendente
from app.models.log import Log
from datetime import datetime, timedelta
from app.utils.notification_utils import format_cadastro_for_notification, broadcast_notification
from app.utils.logger_config import get_logger, log_api_request, log_database_operation
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import os

from reportlab.lib.enums import TA_CENTER

cadastro_bp = Blueprint('cadastro', __name__)
logger = get_logger('cadastro')

@cadastro_bp.route('/index_cadastro')
@login_required
def index_cadastro():
    logger.info(f"Página de cadastro acessada - Usuário: {current_user.username}")
    if current_user.role != 'cadastrador':
        logger.warning(f"Acesso negado à página de cadastro - Usuário: {current_user.username} - Role: {current_user.role}")
        return "Acesso negado!", 403
    municipios = Municipio.query.all()
    estados = Estado.query.all()
    descricoes = Descricao.query.all()
    instituicaoes = Instituicao.query.all()
    atendentes = Atendente.query.all()
    logger.info(f"Dados carregados para página de cadastro - Estados: {len(estados)}, Descrições: {len(descricoes)}, Instituições: {len(instituicaoes)}")
    return render_template('registration_dashboard.html', municipios=municipios, estados=estados, descricoes=descricoes,
                           instituicaoes=instituicaoes, atendentes=atendentes)

@cadastro_bp.route('/cadastro', methods=['POST'])
@login_required
def cadastro():
    logger.info(f"Requisição de cadastro recebida - Usuário: {current_user.username}")
    log_api_request('POST', '/cadastro', current_user.username, 200)
    
    if current_user.role != 'cadastrador':
        logger.warning(f"Acesso negado para cadastro - Usuário: {current_user.username} - Role: {current_user.role}")
        log_api_request('POST', '/cadastro', current_user.username, 403, "Acesso negado")
        return jsonify({'success': False, 'message': 'Acesso negado!'}), 403
    
    try:
        # Log dos dados recebidos
        logger.debug(f"Dados do formulário recebidos - Usuário: {current_user.username}")
        logger.debug(f"Form data: {dict(request.form)}")
        
        nome = request.form.get('nome', '').strip().upper()
        cpf = request.form.get('cpf', '').strip().replace('.', '').replace('-', '')
        telefone = request.form.get('telefone', '').strip().replace('(', '').replace(')', '').replace(' ', '').replace('-', '')
        assentamento = request.form.get('assentamento', '').strip().upper()
        municipio_id = request.form.get('municipio') or None
        estado_id = request.form.get('estado', '').strip()
        descricao_id = request.form.get('descricao', '').strip()
        instituicao_id = request.form.get('instituicao', '').strip()
        
        logger.info(f"Processando cadastro - Nome: {nome}, CPF: {cpf}, Telefone: {telefone}")
        
        # Validação reforçada
        if not all([nome, cpf, telefone, assentamento, estado_id, descricao_id, instituicao_id]):
            logger.warning(f"Campos obrigatórios não preenchidos - Usuário: {current_user.username}")
            log_api_request('POST', '/cadastro', current_user.username, 400, "Campos obrigatórios não preenchidos")
            return jsonify({'success': False, 'message': 'Campos obrigatórios não preenchidos'}), 400
        
        if len(nome) < 3:
            logger.warning(f"Nome muito curto - Usuário: {current_user.username} - Nome: {nome}")
            log_api_request('POST', '/cadastro', current_user.username, 400, "Nome deve ter pelo menos 3 letras")
            return jsonify({'success': False, 'message': 'Nome deve ter pelo menos 3 letras'}), 400
        
        if len(cpf) != 11 or not cpf.isdigit():
            logger.warning(f"CPF inválido - Usuário: {current_user.username} - CPF: {cpf}")
            log_api_request('POST', '/cadastro', current_user.username, 400, "CPF inválido")
            return jsonify({'success': False, 'message': 'CPF inválido'}), 400
        
        if not telefone.isdigit() or len(telefone) < 10:
            logger.warning(f"Telefone inválido - Usuário: {current_user.username} - Telefone: {telefone}")
            log_api_request('POST', '/cadastro', current_user.username, 400, "Telefone inválido")
            return jsonify({'success': False, 'message': 'Telefone inválido'}), 400
        
        if len(assentamento) < 3:
            logger.warning(f"Assentamento inválido - Usuário: {current_user.username} - Assentamento: {assentamento}")
            log_api_request('POST', '/cadastro', current_user.username, 400, "Endereço/Assentamento inválido")
            return jsonify({'success': False, 'message': 'Endereço/Assentamento inválido'}), 400
        
        # Verificar cadastro recente
        dois_minutos_atras = datetime.now() - timedelta(minutes=2)
        cadastro_recente = Cadastro.query.filter(
            Cadastro.cpf == cpf,
            Cadastro.data_hora >= dois_minutos_atras.strftime("%Y-%m-%d %H:%M:%S")
        ).first()
        
        if cadastro_recente:
            logger.warning(f"Cadastro recente encontrado - Usuário: {current_user.username} - CPF: {cpf}")
            log_api_request('POST', '/cadastro', current_user.username, 400, "Cadastro já realizado recentemente")
            return jsonify({
                'success': False,
                'message': 'Cadastro já realizado recentemente para este CPF. Aguarde alguns minutos.'
            }), 400
        
        # Verificar cadastro existente hoje
        hoje_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        hoje_fim = datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        cadastro_existente_hoje = Cadastro.query.filter(
            Cadastro.cpf == cpf,
            Cadastro.descricao_id == descricao_id,
            Cadastro.data_hora >= hoje_inicio.strftime("%Y-%m-%d %H:%M:%S"),
            Cadastro.data_hora <= hoje_fim.strftime("%Y-%m-%d %H:%M:%S")
        ).first()
        
        if cadastro_existente_hoje:
            logger.warning(f"Cadastro existente hoje - Usuário: {current_user.username} - CPF: {cpf}")
            log_api_request('POST', '/cadastro', current_user.username, 400, "Já existe cadastro para este CPF e serviço hoje")
            return jsonify({
                'success': False,
                'message': 'Já existe um cadastro para este CPF e serviço hoje.'
            }), 400
        
        # Criar novo cadastro
        logger.info(f"Criando novo cadastro - Usuário: {current_user.username} - Nome: {nome}")
        novo_cadastro = Cadastro(
            data_hora=datetime.now(),
            nome=nome,
            cpf=cpf,
            telefone=telefone,
            assentamento=assentamento,
            municipio_id=municipio_id,
            estado_id=estado_id,
            descricao_id=descricao_id,
            instituicao_id=instituicao_id,
            atendente_id=current_user.id
        )
        
        # Gerar senha de chamada automaticamente
        novo_cadastro.gerar_senha_chamada()
        
        # Definir prioridade baseada em critérios (pode ser expandido)
        # Por exemplo: se for urgente, prioridade alta
        if descricao_id == '1':  # Assumindo que descrição ID 1 é urgente
            novo_cadastro.definir_prioridade(2, "Cadastro urgente")
        else:
            novo_cadastro.definir_prioridade(0, "Cadastro normal")
        
        # Entrar na fila de espera
        novo_cadastro.entrar_fila()
        
        # Registrar log de criação
        log = Log(
            usuario=current_user.username,
            acao=f"Criou cadastro - Nome: {nome} - Senha: {novo_cadastro.senha_chamada}"
        )
        
        try:
            db.session.add(novo_cadastro)
            db.session.add(log)
            db.session.commit()
            
            logger.info(f"Cadastro criado com sucesso - ID: {novo_cadastro.id} - Senha: {novo_cadastro.senha_chamada}")
            log_database_operation('CREATE', 'cadastro', novo_cadastro.id, current_user.username)
            
            # Redirecionar para o comprovante de senha
            return redirect(url_for('cadastro.senha_comprovante', cadastro_id=novo_cadastro.id))
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Erro ao salvar cadastro - Usuário: {current_user.username} - Erro: {str(e)}")
            return jsonify({'success': False, 'message': 'Erro interno do servidor'}), 500
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erro ao cadastrar - Usuário: {current_user.username} - Erro: {str(e)}")
        log_api_request('POST', '/cadastro', current_user.username, 500, str(e))
        return jsonify({'success': False, 'message': 'Erro interno do servidor. Tente novamente.'}), 500

@cadastro_bp.route('/ver_cadastros')
@login_required
def ver_cadastros():
    if current_user.role not in ['visor', 'admin']:
        return "Acesso negado!"
    nome = request.args.get('nome', '').strip().upper()
    cpf = request.args.get('cpf', '').strip()
    municipio_id = request.args.get('municipio')
    estado_id = request.args.get('estado')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    descricao_id = request.args.get('descricao', '').strip()
    status_filter = request.args.get('status', '').strip()
    cadastros_query = db.session.query(
        Cadastro,
        User.username.label("atendente_nome")
    ).outerjoin(User, Cadastro.atendente_id == User.id)
    if nome:
        cadastros_query = cadastros_query.filter(Cadastro.nome.ilike(f"%{nome}%"))
    if cpf:
        cadastros_query = cadastros_query.filter(Cadastro.cpf.ilike(f"%{cpf}%"))
    if municipio_id and municipio_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.municipio_id == int(municipio_id))
    if estado_id and estado_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.estado_id == int(estado_id))
    if descricao_id and descricao_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.descricao_id == int(descricao_id))
    if status_filter == 'nao_atendidos':
        cadastros_query = cadastros_query.filter(Cadastro.atendida == False)
    elif status_filter == 'atendidos':
        cadastros_query = cadastros_query.filter(Cadastro.atendida == True)
    if data_inicio:
        data_inicio = f"{data_inicio} 00:00:00"
        cadastros_query = cadastros_query.filter(Cadastro.data_hora >= data_inicio)
    if data_fim:
        data_fim = f"{data_fim} 23:59:59"
        cadastros_query = cadastros_query.filter(Cadastro.data_hora <= data_fim)
    cadastros_query = cadastros_query.order_by(Cadastro.data_hora.asc())
    cadastros = cadastros_query.all()
    cadastros_formatados = [
        {
            "cadastro": cadastro,
            "atendente_nome": atendente_nome
        }
        for cadastro, atendente_nome in cadastros
    ]
    estados = Estado.query.all()
    descricoes = Descricao.query.all()
    return render_template('registration_list.html',
                           cadastros=cadastros_formatados,
                           estados=estados,
                           descricoes=descricoes,
                           filtros_ativos={
                               'nome': nome,
                               'cpf': cpf,
                               'municipio': municipio_id,
                               'estado': estado_id,
                               'data_inicio': data_inicio,
                               'data_fim': data_fim,
                               'descricao': descricao_id,
                               'status': status_filter
                           })

@cadastro_bp.route('/export_cadastros')
@login_required
def export_cadastros():
    if current_user.role not in ['admin', 'visor']:
        return "Acesso negado!", 403
    # Filtros iguais ao ver_cadastros
    nome = request.args.get('nome', '').strip().upper()
    cpf = request.args.get('cpf', '').strip()
    municipio_id = request.args.get('municipio')
    estado_id = request.args.get('estado')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    descricao_id = request.args.get('descricao', '').strip()
    status_filter = request.args.get('status', '').strip()
    cadastros_query = db.session.query(
        Cadastro,
        User.username.label("atendente_nome")
    ).outerjoin(User, Cadastro.atendente_id == User.id)
    if nome:
        cadastros_query = cadastros_query.filter(Cadastro.nome.ilike(f"%{nome}%"))
    if cpf:
        cadastros_query = cadastros_query.filter(Cadastro.cpf.ilike(f"%{cpf}%"))
    if municipio_id and municipio_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.municipio_id == int(municipio_id))
    if estado_id and estado_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.estado_id == int(estado_id))
    if descricao_id and descricao_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.descricao_id == int(descricao_id))
    if data_inicio:
        cadastros_query = cadastros_query.filter(Cadastro.data_hora >= f"{data_inicio} 00:00:00")
    if data_fim:
        cadastros_query = cadastros_query.filter(Cadastro.data_hora <= f"{data_fim} 23:59:59")
    if status_filter == 'nao_atendidos':
        cadastros_query = cadastros_query.filter(Cadastro.atendida == False)
    elif status_filter == 'atendidos':
        cadastros_query = cadastros_query.filter(Cadastro.atendida == True)
    cadastros = cadastros_query.all()
    # Gera Excel profissional
    wb = Workbook()
    ws = wb.active
    ws.title = "Cadastros"
    headers = [
        "ID", "Data/Hora", "Nome", "CPF", "Telefone", "Município", "Estado", "Descrição", "Assentamento", "Instituição", "Atendente", "Status"
    ]
    ws.append(headers)
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2c5530")
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    # Excel: bordas, zebra, cabeçalho congelado, fonte Arial, sumário
    from openpyxl.styles.borders import Border, Side
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    zebra_fill = PatternFill("solid", fgColor="F2F2F2")
    arial_font = Font(name='Arial', size=11)
    # Sumário estatístico
    total = len(cadastros)
    atendidos = sum(1 for item, _ in cadastros if item.atendida)
    pendentes = total - atendidos
    ws.insert_rows(1)
    ws['A1'] = f"Total: {total} | Atendidos: {atendidos} | Pendentes: {pendentes}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, color="2c5530", size=13)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws['A3']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = arial_font
            cell.border = thin_border
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = zebra_fill
    # Ajuste de largura e altura (corrigido para MergedCell)
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(15, max_length + 2), 30)
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25
    # Preenche os dados dos cadastros
    for cadastro, atendente_nome in cadastros:
        # Tenta converter data_hora para datetime, se possível
        data_hora_str = cadastro.data_hora
        try:
            data_hora_fmt = datetime.strptime(data_hora_str, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M') if data_hora_str else ''
        except Exception:
            data_hora_fmt = data_hora_str or ''
        ws.append([
            cadastro.id,
            data_hora_fmt,
            cadastro.nome,
            cadastro.cpf,
            cadastro.telefone,
            cadastro.municipio.nome if cadastro.municipio else '',
            cadastro.estado.nome if cadastro.estado else '',
            cadastro.descricao.nome if cadastro.descricao else '',
            cadastro.assentamento,
            cadastro.instituicao.nome if cadastro.instituicao else '',
            atendente_nome or '',
            'Atendido' if cadastro.atendida else 'Pendente'
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="cadastros.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@cadastro_bp.route('/relatorio_cadastros')
@login_required
def relatorio_cadastros():
    if current_user.role not in ['admin', 'visor']:
        return "Acesso negado!", 403
    # Filtros iguais ao ver_cadastros
    nome = request.args.get('nome', '').strip().upper()
    cpf = request.args.get('cpf', '').strip()
    municipio_id = request.args.get('municipio')
    estado_id = request.args.get('estado')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    descricao_id = request.args.get('descricao', '').strip()
    status_filter = request.args.get('status', '').strip()
    cadastros_query = db.session.query(
        Cadastro,
        User.username.label("atendente_nome")
    ).outerjoin(User, Cadastro.atendente_id == User.id)
    if nome:
        cadastros_query = cadastros_query.filter(Cadastro.nome.ilike(f"%{nome}%"))
    if cpf:
        cadastros_query = cadastros_query.filter(Cadastro.cpf.ilike(f"%{cpf}%"))
    if municipio_id and municipio_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.municipio_id == int(municipio_id))
    if estado_id and estado_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.estado_id == int(estado_id))
    if descricao_id and descricao_id.isdigit():
        cadastros_query = cadastros_query.filter(Cadastro.descricao_id == int(descricao_id))
    if data_inicio:
        cadastros_query = cadastros_query.filter(Cadastro.data_hora >= f"{data_inicio} 00:00:00")
    if data_fim:
        cadastros_query = cadastros_query.filter(Cadastro.data_hora <= f"{data_fim} 23:59:59")
    if status_filter == 'nao_atendidos':
        cadastros_query = cadastros_query.filter(Cadastro.atendida == False)
    elif status_filter == 'atendidos':
        cadastros_query = cadastros_query.filter(Cadastro.atendida == True)
    cadastros = cadastros_query.all()
    # PDF: cabeçalho duplo, sumário, linhas zebrada, rodapé institucional
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()
    pstyle = ParagraphStyle('cell', fontName='Helvetica', fontSize=7, leading=9, alignment=0, spaceAfter=1)
    header_style = ParagraphStyle('header', fontName='Helvetica-Bold', fontSize=12, leading=14, alignment=TA_CENTER, textColor=colors.white, backColor=colors.HexColor('#2c5530'), spaceAfter=6)
    # Cabeçalho duplo com logo
    logo_path = os.path.join('app', 'static', 'logo_incra.png')
    if os.path.exists(logo_path):
        img = Image(logo_path, width=2.2*cm, height=2.2*cm)
        img.hAlign = 'CENTER'
        elements.append(img)
    elements.append(Paragraph("<b>MUTIRÃO DA MULHER RURAL</b>", header_style))
    elements.append(Paragraph("<b>Relatório de Cadastros</b>", header_style))
    elements.append(Spacer(1, 0.2*cm))
    # Sumário estatístico
    total = len(cadastros)
    atendidos = sum(1 for item, _ in cadastros if item.atendida)
    pendentes = total - atendidos
    elements.append(Paragraph(f"<b>Total:</b> {total} &nbsp;&nbsp; <b>Atendidos:</b> {atendidos} &nbsp;&nbsp; <b>Pendentes:</b> {pendentes}", styles['Normal']))
    elements.append(Spacer(1, 0.2*cm))
    headers = ["ID", "Data/Hora", "Nome", "CPF", "Telefone", "Município", "Estado", "Descrição", "Assentamento", "Instituição", "Atendente", "Status"]
    data = [[Paragraph(f"<b>{h}</b>", header_style) for h in headers]]
    def trunc(text, maxlen=60):
        return (text[:maxlen] + '...') if text and len(text) > maxlen else (text or '')
    for item, atendente_nome in cadastros:
        row = [
            str(item.id),
            item.data_hora,
            Paragraph(trunc(item.nome, 40), pstyle),
            item.cpf,
            Paragraph(trunc(item.telefone, 20), pstyle),
            Paragraph(trunc(item.municipio.nome if item.municipio else '', 30), pstyle),
            Paragraph(trunc(item.estado.nome if item.estado else '', 20), pstyle),
            Paragraph(trunc(item.descricao.nome if item.descricao else '', 30), pstyle),
            Paragraph(trunc(item.assentamento, 30), pstyle),
            Paragraph(trunc(item.instituicao.nome if item.instituicao else '', 30), pstyle),
            Paragraph(trunc(atendente_nome, 25), pstyle),
            Paragraph('Atendida' if item.atendida else 'Pendente', pstyle)
        ]
        data.append(row)
    col_widths = [1.2*cm, 2.5*cm, 3.2*cm, 2.2*cm, 2.2*cm, 2.5*cm, 2.2*cm, 2.7*cm, 2.7*cm, 2.7*cm, 2.7*cm, 2*cm]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c5530')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightgrey]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (0,1), (-1,-1), 'LEFT'),
        ('FONTSIZE', (0,1), (-1,-1), 7),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
    ]))
    elements.append(table)
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.drawString(1*cm, 1*cm, f"Mutirão da Mulher Rural - Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        canvas.drawRightString(landscape(A4)[0]-1*cm, 1*cm, f"Página {doc.page}")
        canvas.restoreState()
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="relatorio_cadastros.pdf", mimetype="application/pdf")

@cadastro_bp.route('/toggle_atendimento/<int:id>', methods=['POST'])
@login_required
def toggle_atendimento(id):
    if current_user.role not in ['admin', 'visor']:
        return jsonify({'success': False, 'message': 'Acesso negado!'}), 403
    cadastro = Cadastro.query.get(id)
    if not cadastro:
        return jsonify({'success': False, 'message': 'Cadastro não encontrado!'}), 404
    try:
        cadastro.atendida = not bool(cadastro.atendida)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Status alterado com sucesso!', 'atendida': cadastro.atendida})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro ao alterar status: {str(e)}'})

@cadastro_bp.route('/chamar/<int:cadastro_id>', methods=['POST', 'GET'])
@login_required
def chamar_beneficiaria(cadastro_id):
    if current_user.role not in ['visor', 'admin']:
        flash('Acesso negado!', 'error')
        return redirect(url_for('cadastro.ver_cadastros'))
    cadastro = Cadastro.query.get_or_404(cadastro_id)
    from datetime import datetime
    # Atualiza status e horário
    cadastro.status_chamada = 'chamado'
    cadastro.horario_chamada = datetime.now()
    cadastro.registrar_chamada(current_user.username)
    db.session.commit()
    return render_template('chamada_painel.html', cadastro=cadastro)

@cadastro_bp.route('/marcar_atendido/<int:cadastro_id>', methods=['POST'])
@login_required
def marcar_atendido(cadastro_id):
    if current_user.role not in ['visor', 'admin']:
        return jsonify({'success': False, 'message': 'Acesso negado!'}), 403
    try:
        cadastro = Cadastro.query.get_or_404(cadastro_id)
        cadastro.status_chamada = 'atendido'
        cadastro.atendida = True
        cadastro.registrar_chamada(current_user.username)
        db.session.commit()
        
        # Registrar log
        log = Log(usuario=current_user.username, acao=f"Marcou como atendido: {cadastro.nome}")
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Beneficiária marcada como atendida!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@cadastro_bp.route('/marcar_ausente/<int:cadastro_id>', methods=['POST'])
@login_required
def marcar_ausente(cadastro_id):
    if current_user.role not in ['visor', 'admin']:
        return jsonify({'success': False, 'message': 'Acesso negado!'}), 403
    try:
        cadastro = Cadastro.query.get_or_404(cadastro_id)
        cadastro.status_chamada = 'ausente'
        cadastro.registrar_chamada(current_user.username)
        db.session.commit()
        
        # Registrar log
        log = Log(usuario=current_user.username, acao=f"Marcou como ausente: {cadastro.nome}")
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Beneficiária marcada como ausente!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@cadastro_bp.route('/historico_chamadas/<int:cadastro_id>')
@login_required
def historico_chamadas(cadastro_id):
    if current_user.role not in ['visor', 'admin']:
        return "Acesso negado!", 403
    cadastro = Cadastro.query.get_or_404(cadastro_id)
    
    # Parse do histórico de chamadas
    historico = []
    if cadastro.historico_chamadas:
        import json
        try:
            historico = json.loads(cadastro.historico_chamadas)
        except:
            historico = []
    
    return render_template('historico_chamadas.html', cadastro=cadastro, historico=historico)

@cadastro_bp.route('/proximo_fila', methods=['POST'])
@login_required
def proximo_fila():
    if current_user.role not in ['visor', 'admin']:
        return jsonify({'success': False, 'message': 'Acesso negado!'}), 403
    
    try:
        # Busca o próximo da fila por prioridade
        proximo = Cadastro.query.filter(
            Cadastro.status_chamada == 'aguardando',
            Cadastro.posicao_fila.isnot(None)
        ).order_by(
            Cadastro.prioridade.desc(),
            Cadastro.posicao_fila.asc()
        ).first()
        
        if proximo:
            return jsonify({
                'success': True,
                'cadastro': {
                    'id': proximo.id,
                    'nome': proximo.nome,
                    'senha': proximo.senha_chamada,
                    'prioridade': proximo.prioridade,
                    'posicao': proximo.posicao_fila
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Nenhuma pessoa na fila!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@cadastro_bp.route('/definir_prioridade/<int:cadastro_id>', methods=['POST'])
@login_required
def definir_prioridade(cadastro_id):
    if current_user.role not in ['visor', 'admin']:
        return jsonify({'success': False, 'message': 'Acesso negado!'}), 403
    
    try:
        data = request.get_json()
        nivel = data.get('nivel', 0)
        
        cadastro = Cadastro.query.get_or_404(cadastro_id)
        cadastro.definir_prioridade(nivel)
        cadastro.registrar_chamada(current_user.username, f"definiu prioridade {nivel}")
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Prioridade definida como {["Normal", "Alta", "Urgente"][nivel]}!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@cadastro_bp.route('/entrar_fila/<int:cadastro_id>', methods=['POST'])
@login_required
def entrar_fila(cadastro_id):
    if current_user.role not in ['visor', 'admin']:
        return jsonify({'success': False, 'message': 'Acesso negado!'}), 403
    
    try:
        cadastro = Cadastro.query.get_or_404(cadastro_id)
        cadastro.entrar_fila()
        cadastro.registrar_chamada(current_user.username, "entrou na fila")
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Cadastro colocado na fila na posição {cadastro.posicao_fila}!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@cadastro_bp.route('/sair_fila/<int:cadastro_id>', methods=['POST'])
@login_required
def sair_fila(cadastro_id):
    if current_user.role not in ['visor', 'admin']:
        return jsonify({'success': False, 'message': 'Acesso negado!'}), 403
    
    try:
        cadastro = Cadastro.query.get_or_404(cadastro_id)
        cadastro.sair_fila()
        cadastro.registrar_chamada(current_user.username, "saiu da fila")
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Cadastro removido da fila!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500

@cadastro_bp.route('/fila_chamadas')
@login_required
def fila_chamadas():
    if current_user.role not in ['visor', 'admin']:
        return "Acesso negado!", 403
    
    # Busca todos os cadastros na fila
    fila = Cadastro.query.filter(
        Cadastro.status_chamada == 'aguardando',
        Cadastro.posicao_fila.isnot(None)
    ).order_by(
        Cadastro.prioridade.desc(),
        Cadastro.posicao_fila.asc()
    ).all()
    
    return render_template('fila_chamadas.html', fila=fila)

@cadastro_bp.route('/senha/<int:cadastro_id>/comprovante')
@login_required
def senha_comprovante(cadastro_id):
    cadastro = Cadastro.query.get_or_404(cadastro_id)
    data = cadastro.data_hora.strftime('%d/%m/%Y') if cadastro.data_hora else '--'
    hora = cadastro.data_hora.strftime('%H:%M') if cadastro.data_hora else '--'
    servico = getattr(cadastro.descricao, 'nome', '--') if cadastro.descricao else '--'
    guiche = getattr(cadastro.atendente, 'username', '--') if cadastro.atendente else '--'
    senha = cadastro.senha_chamada or '--'
    nome = cadastro.nome or '--'
    return render_template(
        'senha_comprovante.html',
        senha=senha,
        nome=nome,
        data=data,
        hora=hora,
        servico=servico,
        guiche=guiche
    )
